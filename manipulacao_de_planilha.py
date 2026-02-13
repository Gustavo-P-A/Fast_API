from openpyxl import load_workbook, Workbook

def criar_aba(categoria, planilha_principal):
    if categoria not in planilha_principal.sheetnames:
        planilha_principal.create_sheet(categoria)
        nova = planilha_principal[categoria]
        # Define os cabeçalhos da nova aba
        nova['A1'].value = 'data'
        nova['B1'].value = 'categoria'
        nova['C1'].value = 'descricao'
        nova['D1'].value = 'valor'

def transferir_aba(aba_origem, aba_destino, linha_origem):
    linha_destino = aba_destino.max_row + 1
    for col in range(1, 5):
        valor = aba_origem.cell(row=linha_origem, column=col).value
        aba_destino.cell(row=linha_destino, column=col).value = valor

# Carrega o arquivo
try:
    planilha_principal = load_workbook('testexls.xlsx')
    aba_principal = planilha_principal['Sheet1']
except FileNotFoundError:
    print("Erro: O arquivo 'testexls.xlsx' não foi encontrado.")
    exit()

# Itera sobre as linhas da planilha principal (começando da 2)
ultima_linha = aba_principal.max_row

for linha in range(2, ultima_linha + 1):
    categoria = aba_principal.cell(row=linha, column=2).value
    
    # Se a célula de categoria estiver vazia, interrompe o loop
    if not categoria:
        continue
    
    # Cria a aba se não existir
    criar_aba(str(categoria), planilha_principal)
    
    # Define o destino e transfere os dados
    aba_destino = planilha_principal[str(categoria)]
    transferir_aba(aba_principal, aba_destino, linha)

# Salva as alterações
planilha_principal.save('testexls1.xlsx')
print("Processo concluído e arquivo salvo!")

