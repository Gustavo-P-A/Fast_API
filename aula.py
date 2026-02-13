from openpyxl import load_workbook

arquivo = load_workbook("testexls.xlsx")

#ver abas
print(arquivo.sheetnames)

#pegar abas ativa
aba_atual = arquivo.active
print(aba_atual)

# selecionar abas especificas
aba_atual = arquivo['Sheet1']
print(aba_atual)

# selecionar celulas
val1= aba_atual['B3'].value
val2 = aba_atual.cell(row=3, column=2).value # a funcao cell permite pergar a palavra q esta na linha 1 coluna 3
print(val1)
print(val2) 

#editar 
# aba_atual.cell(row=1, column=1).value = "daude"
# arquivo.save('testexls.xlsx') # posso criar outro arquivo xlsx ou subistituir

#ultima linha
print(aba_atual.max_row) #pode acabar contando a mais
print(aba_atual.max_column)
print(len(aba_atual['A']))

